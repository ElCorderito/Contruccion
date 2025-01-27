from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic.edit import CreateView
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.db import transaction
from .models import *
from .forms import *
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
import json
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.core.paginator import Paginator
from datetime import datetime
from django.views.generic import ListView
from django.views.generic import TemplateView
from django.views import View
from django.views.generic import DetailView
from django.views.generic import UpdateView
from django.views.generic import DeleteView
from django.db.models import Q
from django.core.exceptions import PermissionDenied
from django.http import Http404
from django.core.mail import send_mail
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


## Login ##


def landing(request):
    return render(request, 'landing.html')


### Inicio ###


class HomeView(TemplateView):
    template_name = 'inicio.html'


### Proyecto ###


class ProyectoCreateView(CreateView):
    model = Proyecto
    form_class = ProyectoForm
    template_name = 'crear_proyecto.html'
    success_url = reverse_lazy('proyecto_lista')

    def form_valid(self, form):
        # Si el formulario es válido, sigue el flujo normal
        return super().form_valid(form)

    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class CambiarEstadoProyectoView(View):
    def post(self, request, pk):
        proyecto = Proyecto.objects.get(pk=pk)
        proyecto.estado = 'terminado'
        proyecto.save()
        return redirect(reverse('proyecto_lista'))


class ProyectoListaView(TemplateView):
    template_name = 'proyecto_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyectos_en_proceso'] = Proyecto.objects.filter(estado='en_proceso')
        context['proyectos_terminados'] = Proyecto.objects.filter(estado='terminado')
        return context


def proyecto_detalle(request, pk):
    # Obtener el proyecto
    proyecto = get_object_or_404(Proyecto, pk=pk)
    
    # Obtener el presupuesto asociado, si lo hay
    presupuesto = Presupuesto.objects.filter(proyecto=proyecto).first()
    monto_restante = presupuesto.calcular_monto_restante() if presupuesto else 0
    
    # Obtener los documentos asociados al proyecto (sin paginación)
    documentos = DocumentoProyecto.objects.filter(proyecto=proyecto)
    
    # Obtener los inventarios asociados al proyecto con paginación
    inventarios_list = Inventario.objects.filter(proyecto=proyecto).order_by('fecha')
    
    paginator = Paginator(inventarios_list, 10)  # Paginación de 10 inventarios por página
    page_number = request.GET.get('page')
    inventarios = paginator.get_page(page_number)
    
    # Renderizar la plantilla con los datos del proyecto
    context = {
        'proyecto': proyecto,
        'presupuesto': presupuesto,
        'monto_restante': monto_restante,
        'documentos': documentos,
        'inventarios': inventarios  # Solo los inventarios tienen paginación
    }
    return render(request, 'proyecto_detalle.html', context)


class ProyectoConfirmDeleteView(DetailView):
    model = Proyecto
    template_name = 'proyecto_confirm_delete.html'
    context_object_name = 'proyecto'


class ProyectoDeleteView(DeleteView):
    model = Proyecto
    template_name = 'proyecto_delete.html'  # (Podrías no usarlo si vas 100% con modal)
    success_url = reverse_lazy('proyecto_lista')


### Categoria ###


class CategoriaCreateView(CreateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'crear_categoria.html'
    success_url = reverse_lazy('categoria_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class CategoriaListaView(TemplateView):
    template_name = 'categoria_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = Categoria.objects.all()
        return context


class CategoriaUpdateView(UpdateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'categoria_editar.html'
    success_url = reverse_lazy('categoria_lista')


class CategoriaDeleteView(DeleteView):
    model = Categoria
    success_url = reverse_lazy('categoria_lista')

### Clasificación ###


class ClasificacionCreateView(CreateView):
    model = Clasificacion
    form_class = ClasificacionForm
    template_name = 'crear_clasificacion.html'
    success_url = reverse_lazy('clasificacion_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class ClasificacionListaView(TemplateView):
    template_name = 'clasificacion_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['clasificaciones'] = Clasificacion.objects.all()
        return context


class ClasificacionUpdateView(UpdateView):
    model = Clasificacion
    form_class = ClasificacionForm
    template_name = 'clasificacion_editar.html'
    success_url = reverse_lazy('clasificacion_lista')


class ClasificacionDeleteView(DeleteView):
    model = Clasificacion
    success_url = reverse_lazy('clasificacion_lista')


### Material ###


def materiales(request):
    query = request.GET.get('q')  # Obtener el término de búsqueda de la barra de búsqueda
    if query:
        materiales_list = Material.objects.filter(nombre__icontains=query).order_by('nombre')
    else:
        materiales_list = Material.objects.all().order_by('nombre')  # Obtenemos todos los materiales y los ordenamos
    
    paginator = Paginator(materiales_list, 10)  # Creamos un Paginator, 10 materiales por página

    page_number = request.GET.get('page')  # Obtenemos el número de página de los parámetros de la URL
    materiales = paginator.get_page(page_number)  # Obtenemos la página correspondiente

    return render(request, "material_lista.html", {'materiales': materiales, 'q': query})


class MaterialUpdateView(UpdateView):
    model = Material
    form_class = MaterialForm
    template_name = 'material_editar.html'
    success_url = reverse_lazy('material_lista')


class MaterialCreateView(CreateView):
    model = Material
    form_class = MaterialForm
    template_name = 'crear_material.html'
    success_url = reverse_lazy('material_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class MaterialDeleteView(DeleteView):
    model = Material
    success_url = reverse_lazy('material_lista')


### Tienda ###


class TiendaListaView(TemplateView):
    template_name = 'tienda_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tiendas'] = Tienda.objects.all()
        return context


class TiendaUpdateView(UpdateView):
    model = Tienda
    form_class = TiendaForm
    template_name = 'tienda_editar.html'
    success_url = reverse_lazy('tienda_lista')


class TiendaCreateView(CreateView):
    model = Tienda
    form_class = TiendaForm
    template_name = 'crear_tienda.html'
    success_url = reverse_lazy('tienda_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class TiendaDeleteView(DeleteView):
    model = Tienda
    success_url = reverse_lazy('tienda_lista')

### Proveedores ###


class ProveedorListaView(TemplateView):
    template_name = 'proveedor_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proveedores'] = Proveedor.objects.all()
        return context


class ProveedorUpdateView(UpdateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'proveedor_editar.html'
    success_url = reverse_lazy('proveedor_lista')


class ProveedorCreateView(CreateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'crear_proveedor.html'
    success_url = reverse_lazy('proveedor_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class ProveedorDeleteView(DeleteView):
    model = Proveedor
    success_url = reverse_lazy('proveedor_lista')


### Fase ###


class FaseListaView(TemplateView):
    template_name = 'fase_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fases'] = Fase.objects.all()
        return context


class FaseUpdateView(UpdateView):
    model = Fase
    form_class = FaseForm
    template_name = 'fase_editar.html'
    success_url = reverse_lazy('fase_lista')


class FaseCreateView(CreateView):
    model = Fase
    form_class = FaseForm
    template_name = 'crear_fase.html'
    success_url = reverse_lazy('fase_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


### Etapa ###


class EtapaListaView(TemplateView):
    template_name = 'etapa_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['etapas'] = Etapa.objects.all()
        return context


class EtapaUpdateView(UpdateView):
    model = Etapa
    form_class = EtapaForm
    template_name = 'etapa_editar.html'
    success_url = reverse_lazy('etapa_lista')


class EtapaCreateView(CreateView):
    model = Etapa
    form_class = EtapaForm
    template_name = 'crear_etapa.html'
    success_url = reverse_lazy('etapa_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


### Inventario ###


class InventarioListaView(TemplateView):
    template_name = 'inventario_lista.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Filtrar los proyectos que están en proceso
        proyectos_en_proceso = Proyecto.objects.filter(estado='en_proceso')
        # Filtrar inventarios relacionados con esos proyectos
        inventarios = Inventario.objects.filter(proyecto__in=proyectos_en_proceso)
        context['inventarios_por_proyecto'] = self.agrupar_por_proyecto(inventarios)
        return context

    def agrupar_por_proyecto(self, inventarios):
        inventarios_por_proyecto = {}
        for inventario in inventarios:
            proyecto = inventario.proyecto
            if proyecto not in inventarios_por_proyecto:
                inventarios_por_proyecto[proyecto] = []
            inventarios_por_proyecto[proyecto].append(inventario)
        return inventarios_por_proyecto


class InventarioCreateView(CreateView):
    model = Inventario
    form_class = InventarioForm
    template_name = 'crear_inventario.html'
    success_url = reverse_lazy('inventario_lista')

    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class InventarioUpdateView(UpdateView):
    model = Inventario
    form_class = InventarioForm
    template_name = 'inventario_editar.html'
    success_url = reverse_lazy('inventario_lista')


class InventarioDeleteView(DeleteView):
    model = Inventario
    success_url = reverse_lazy('inventario_lista')


## Presupuesto ##


class ProyectoDetalleView(DetailView):
    model = Proyecto
    template_name = 'proyecto_detalle.html'
    context_object_name = 'proyecto_lista'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proyecto = self.get_object()

        # Obtener el presupuesto asociado al proyecto
        presupuesto = Presupuesto.objects.filter(proyecto=proyecto).first()
        context['presupuesto'] = presupuesto
        
        if presupuesto:
            # Calcular el total gastado en el inventario
            inventarios = Inventario.objects.filter(proyecto=proyecto)
            total_gastado_inventario = inventarios.aggregate(total=Sum('subtotal'))['total'] or 0

            # Actualizar el monto gastado en el presupuesto
            presupuesto.monto_gastado = total_gastado_inventario
            presupuesto.save()

            # Calcular el monto restante
            context['monto_restante'] = presupuesto.calcular_monto_restante()

        # Añadir inventarios al contexto
        context['inventarios'] = Inventario.objects.filter(proyecto=proyecto)
        return context
    

class PresupuestoCreateView(CreateView):
    model = Presupuesto
    form_class = PresupuestoForm
    template_name = 'crear_presupuesto.html'
    
    def dispatch(self, request, *args, **kwargs):
        proyecto = Proyecto.objects.get(pk=self.kwargs['pk'])
        if hasattr(proyecto, 'presupuesto'):
            # Redirigir si el presupuesto ya existe
            return redirect('proyecto_detalle', pk=proyecto.pk)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.proyecto = Proyecto.objects.get(pk=self.kwargs['pk'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('proyecto_detalle', kwargs={'pk': self.kwargs['pk']})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pk'] = self.kwargs['pk']  # Pasar el pk al contexto
        return context

    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))


class PresupuestoUpdateView(UpdateView):
    model = Presupuesto
    form_class = PresupuestoForm
    template_name = 'presupuesto_editar.html'
    
    def get_success_url(self):
        return reverse_lazy('proyecto_detalle', kwargs={'pk': self.object.proyecto.pk})


## Documentos ##


class DocumentoProyectoCreateView(CreateView):
    model = DocumentoProyecto
    form_class = DocumentoProyectoForm
    template_name = 'subir_documento.html'

    def form_valid(self, form):
        # Obtener el proyecto desde la URL y asignarlo al documento
        form.instance.proyecto = get_object_or_404(Proyecto, pk=self.kwargs['pk'])
        return super().form_valid(form)

    def get_success_url(self):
        # Redirigir a la página de detalles del proyecto después de guardar el documento
        return reverse_lazy('proyecto_detalle', kwargs={'pk': self.kwargs['pk']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasar el proyecto al contexto por si es necesario en la plantilla
        context['proyecto_pk'] = self.kwargs['pk']
        return context
    
    def form_invalid(self, form):
        # Si el formulario no es válido, agrega un mensaje de error al contexto
        error_message = "Hay errores en el formulario. Por favor, corrígelos."
        return self.render_to_response(self.get_context_data(form=form, error_message=error_message))
    

class TodosDocumentosListView(ListView):
    model = DocumentoProyecto
    template_name = 'todos_documentos.html'
    context_object_name = 'documentos'

    def get_queryset(self):
        # Devuelve todos los documentos en la base de datos
        return DocumentoProyecto.objects.all()


## Reportes ##


def generar_reporte(request):
    form = ReporteForm(request.GET or None)
    inventarios = Inventario.objects.all()

    if form.is_valid():
        proyecto = form.cleaned_data.get('proyecto')
        fecha_inicio = form.cleaned_data.get('fecha_inicio')
        fecha_fin = form.cleaned_data.get('fecha_fin')
        categoria = form.cleaned_data.get('categoria')
        clasificacion = form.cleaned_data.get('clasificacion')
        tienda = form.cleaned_data.get('tienda')
        proveedor = form.cleaned_data.get('proveedor')
        etapa_seleccionada = form.cleaned_data.get('etapa')  # <-- si quieres filtrar por una etapa en específico
        material_search = form.cleaned_data.get('material_search')

        if proyecto:
            inventarios = inventarios.filter(proyecto=proyecto)
        if fecha_inicio and fecha_fin:
            inventarios = inventarios.filter(fecha__range=[fecha_inicio, fecha_fin])
        if categoria:
            inventarios = inventarios.filter(material__categoria=categoria)
        if clasificacion:
            inventarios = inventarios.filter(material__clasificacion=clasificacion)
        if tienda:
            inventarios = inventarios.filter(tienda=tienda)
        if proveedor:
            inventarios = inventarios.filter(proveedor=proveedor)
        if etapa_seleccionada:
            inventarios = inventarios.filter(etapa=etapa_seleccionada)
        if material_search:
            inventarios = inventarios.filter(material__nombre__icontains=material_search)

    total_gasto = inventarios.aggregate(total=Sum('subtotal'))['total'] or 0

    # Si deseas obtener TODAS las etapas existentes
    # (si siempre son 3 fijas, puedes filtrar manualmente)
    etapas = Etapa.objects.all().order_by('id')  # Ajusta el .order_by si quieres un orden particular

    # Crear una lista/diccionario que relacione "etapa" con los inventarios que se usaron en esa etapa
    inventarios_por_etapa = []
    for e in etapas:
        inv_etapa = inventarios.filter(etapa=e)
        inventarios_por_etapa.append({
            'etapa': e,            # El objeto Etapa
            'inventarios': inv_etapa  # El queryset filtrado
        })

    # PREVISUALIZACIÓN
    if request.GET.get('preview'):
        return render(request, 'reporte_previa.html', {
            'form': form,
            'inventarios_por_etapa': inventarios_por_etapa,
            'total_gasto': total_gasto,
        })

    # GENERAR PDF
    if request.GET.get('generar_pdf'):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'

        p = canvas.Canvas(response, pagesize=letter)
        p.drawString(100, 750, "Reporte de Gastos")

        y = 700
        # Aquí, si quieres que el PDF también vaya etapa por etapa
        for data in inventarios_por_etapa:
            etapa = data['etapa']
            etapa_inventarios = data['inventarios']

            # Título de la Etapa
            p.drawString(100, y, f"Etapa: {etapa.nombre}")
            y -= 20

            if not etapa_inventarios.exists():
                p.drawString(100, y, "No se usaron materiales en esta etapa")
                y -= 20
            else:
                for inv in etapa_inventarios:
                    text_line = (
                        f"Material: {inv.material.nombre}, "
                        f"Cantidad: {inv.cantidad}, "
                        f"Subtotal: {inv.subtotal}$"
                    )
                    p.drawString(100, y, text_line)
                    y -= 20
                    if y < 50:
                        p.showPage()
                        y = 750

            # Espacio entre etapas
            y -= 20
            if y < 50:
                p.showPage()
                y = 750

        # Total general
        p.drawString(100, 50, f"Total Gasto: {total_gasto}$")
        p.showPage()
        p.save()

        return response
    
    # GENERAR EXCEL (nuevo)
    if request.GET.get('generar_excel'):
        return generar_excel_response(inventarios_por_etapa, total_gasto)

    # SI NO SE HA SELECCIONADO NI preview NI generar_pdf
    return render(request, 'reporte_filtros.html', {
        'form': form
    })


def generar_excel_response(inventarios_por_etapa, total_gasto):
    """Genera el archivo Excel en memoria y lo retorna en un HttpResponse."""
    # 1. Crear el Workbook
    wb = Workbook()

    # Por defecto openpyxl crea una hoja llamada "Sheet"; podemos usarla o eliminarla.
    # Vamos a usarla para un posible resumen general. O la quitamos y creamos siempre nuevas.
    wb.remove(wb.active)

    # 2. Para cada etapa, creamos una hoja nueva
    for data in inventarios_por_etapa:
        etapa = data['etapa']
        invs = data['inventarios']

        # Nombrar la hoja (máximo 31 caracteres en Excel)
        sheet_name = (etapa.nombre[:31]) if etapa.nombre else "Etapa"
        ws = wb.create_sheet(title=sheet_name)

        # Encabezado de columnas
        headers = ["Material", "Cantidad", "Precio Unitario", "Subtotal", "Fecha", "Tienda", "Proveedor"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Contenido
        row_num = 2
        for inv in invs:
            ws.cell(row=row_num, column=1, value=inv.material.nombre)
            ws.cell(row=row_num, column=2, value=inv.cantidad)
            ws.cell(row=row_num, column=3, value=inv.precio_unitario)
            ws.cell(row=row_num, column=4, value=inv.subtotal)
            ws.cell(row=row_num, column=5, value=str(inv.fecha))  # Convertir a string
            ws.cell(row=row_num, column=6, value=inv.tienda.nombre if inv.tienda else "-")
            ws.cell(row=row_num, column=7, value=inv.proveedor.nombre if inv.proveedor else "-")
            row_num += 1

        # Si no se usaron materiales, puedes poner un mensaje
        if not invs.exists():
            ws.cell(row=2, column=1, value="No se usaron materiales en esta etapa")

    # OPCIONAL: Crear una hoja resumen general
    ws_summary = wb.create_sheet(title="Resumen")
    ws_summary["A1"] = "Total Gasto"
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["B1"] = total_gasto

    # 3. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte.xlsx"'

    # 4. Guardar el workbook en la respuesta
    wb.save(response)
    return response
